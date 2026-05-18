from __future__ import annotations

import asyncio
import io
import json
import time
from dataclasses import dataclass, asdict, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Deque, List, Optional
from collections import deque

from fastapi import FastAPI, WebSocket, WebSocketDisconnect
from fastapi.responses import HTMLResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from openpyxl import Workbook
from pydantic import BaseModel

from .sources import EEGSource, StreamInfo, get_source
from .ws_manager import ConnectionManager


ROOT_DIR = Path(__file__).resolve().parent.parent
FRONTEND_DIR = ROOT_DIR / "frontend"

app = FastAPI(title="X.on EEG Live POC", version="0.2.0")
app.mount("/static", StaticFiles(directory=str(FRONTEND_DIR), html=False), name="static")


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="milliseconds")


@dataclass
class Marker:
    timestamp: str
    label: str


@dataclass
class SessionState:
    running: bool = False
    started_at: Optional[str] = None
    started_monotonic: Optional[float] = None
    markers: List[Marker] = field(default_factory=list)
    chunks: List[dict] = field(default_factory=list)  # {"t0": iso, "values": [[..]]}
    sample_count: int = 0

    def duration_seconds(self) -> float:
        if self.started_monotonic is None:
            return 0.0
        return round(time.monotonic() - self.started_monotonic, 2)


class AppState:
    def __init__(self) -> None:
        self.source: Optional[EEGSource] = None
        self.info: Optional[StreamInfo] = None
        self.session = SessionState()
        self.last_chunks: Deque[dict] = deque(maxlen=400)
        self.last_impedance: Optional[dict] = None
        self.lock = asyncio.Lock()


state = AppState()


manager = ConnectionManager()


async def on_chunk(t0: str, values: List[List[float]]) -> None:
    chunk = {"t0": t0, "values": values}
    state.last_chunks.append(chunk)
    if state.session.running:
        state.session.chunks.append(chunk)
        state.session.sample_count += len(values)
    await manager.broadcast({
        "type": "eeg_chunk",
        "fs": state.info.fs if state.info else None,
        "t0": t0,
        "values": values,
    })


async def on_impedance(t: str, values_kohm: List[Optional[float]]) -> None:
    payload = {"type": "impedance", "t": t, "values_kohm": values_kohm}
    state.last_impedance = payload
    await manager.broadcast(payload)


@app.on_event("startup")
async def startup_event() -> None:
    state.source = get_source()
    state.info = state.source.info()
    await state.source.start(on_chunk, on_impedance)


@app.on_event("shutdown")
async def shutdown_event() -> None:
    if state.source is not None:
        await state.source.stop()


@app.get("/", response_class=HTMLResponse)
async def index() -> HTMLResponse:
    return HTMLResponse((FRONTEND_DIR / "index.html").read_text(encoding="utf-8"))


@app.get("/api/stream/info")
async def stream_info() -> dict:
    if state.info is None:
        return {}
    return {
        "channels": state.info.channels,
        "fs": state.info.fs,
        "reference": state.info.reference,
        "source": state.info.source,
        "units": state.info.units,
        "session": {
            "running": state.session.running,
            "started_at": state.session.started_at,
            "duration_s": state.session.duration_seconds(),
            "sample_count": state.session.sample_count,
            "marker_count": len(state.session.markers),
        },
    }


class MarkerBody(BaseModel):
    label: Optional[str] = None


@app.post("/api/session/start")
async def session_start() -> dict:
    async with state.lock:
        state.session = SessionState(
            running=True,
            started_at=now_iso(),
            started_monotonic=time.monotonic(),
        )
    msg = {
        "type": "session",
        "state": "running",
        "started_at": state.session.started_at,
        "duration_s": 0.0,
    }
    await manager.broadcast(msg)
    return msg


@app.post("/api/session/stop")
async def session_stop() -> dict:
    async with state.lock:
        state.session.running = False
        duration = state.session.duration_seconds()
        summary = {
            "type": "session",
            "state": "idle",
            "started_at": state.session.started_at,
            "duration_s": duration,
            "sample_count": state.session.sample_count,
            "marker_count": len(state.session.markers),
        }
    await manager.broadcast(summary)
    return summary


@app.post("/api/session/marker")
async def session_marker(body: MarkerBody) -> dict:
    label = (body.label or "marker").strip() or "marker"
    marker = Marker(timestamp=now_iso(), label=label)
    async with state.lock:
        state.session.markers.append(marker)
    payload = {"type": "marker", "t": marker.timestamp, "label": marker.label}
    await manager.broadcast(payload)
    return payload


@app.post("/api/session/reset")
async def session_reset() -> dict:
    async with state.lock:
        state.session = SessionState()
        state.last_chunks.clear()
    return {"ok": True}


@app.post("/api/impedance/start")
async def impedance_start() -> dict:
    if state.source is not None:
        await state.source.set_impedance_streaming(True)
    return {"ok": True, "enabled": True}


@app.post("/api/impedance/stop")
async def impedance_stop() -> dict:
    if state.source is not None:
        await state.source.set_impedance_streaming(False)
    return {"ok": True, "enabled": False}


@app.get("/api/export/excel")
async def export_excel() -> StreamingResponse:
    info = state.info
    channels = info.channels if info else []
    fs = info.fs if info else None

    chunks = list(state.session.chunks) if state.session.chunks else list(state.last_chunks)
    markers = list(state.session.markers)

    workbook = Workbook()
    eeg_sheet = workbook.active
    eeg_sheet.title = "EEG"
    header = ["timestamp_chunk_t0", "sample_idx_in_chunk", *channels]
    eeg_sheet.append(header)
    for chunk in chunks:
        t0 = chunk["t0"]
        for idx, row in enumerate(chunk["values"]):
            padded = list(row) + [None] * max(0, len(channels) - len(row))
            eeg_sheet.append([t0, idx, *padded[: len(channels)]])

    mk_sheet = workbook.create_sheet("Markers")
    mk_sheet.append(["timestamp", "label"])
    for m in markers:
        mk_sheet.append([m.timestamp, m.label])

    meta_sheet = workbook.create_sheet("Meta")
    meta_sheet.append(["key", "value"])
    if info:
        meta_sheet.append(["source", info.source])
        meta_sheet.append(["fs_hz", fs])
        meta_sheet.append(["reference", info.reference])
        meta_sheet.append(["units", info.units])
        meta_sheet.append(["channels", ", ".join(channels)])
    meta_sheet.append(["session_started_at", state.session.started_at or ""])
    meta_sheet.append(["session_duration_s", state.session.duration_seconds()])
    meta_sheet.append(["session_sample_count", state.session.sample_count])

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    filename = f"xon_eeg_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@app.websocket("/ws/eeg")
async def eeg_websocket(websocket: WebSocket) -> None:
    await manager.connect(websocket)
    try:
        if state.info:
            await websocket.send_text(json.dumps({
                "type": "info",
                "channels": state.info.channels,
                "fs": state.info.fs,
                "reference": state.info.reference,
                "source": state.info.source,
                "units": state.info.units,
            }))
        if state.last_impedance:
            await websocket.send_text(json.dumps(state.last_impedance))
        if state.session.running:
            await websocket.send_text(json.dumps({
                "type": "session",
                "state": "running",
                "started_at": state.session.started_at,
                "duration_s": state.session.duration_seconds(),
            }))
        while True:
            await websocket.receive_text()
    except WebSocketDisconnect:
        await manager.disconnect(websocket)
    except Exception:
        await manager.disconnect(websocket)
